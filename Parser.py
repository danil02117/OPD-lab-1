import requests
from bs4 import BeautifulSoup
import openpyxl

def Parser():
    url = 'https://www.labirint.ru/books/'

    response = requests.get(url)

    soup = BeautifulSoup(response.content, 'html.parser')

    books = soup.find_all('div', {'class': 'product'})

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value='Book')
    worksheet.cell(row=1, column=2, value='Author')
    worksheet.cell(row=1, column=3, value='Price')

    for i, book in enumerate(books, start=2):
        title_elem = book.find('a', {'class': 'cover'})
        title = title_elem['title'].strip() if title_elem else ''

        author_elem = book.find('a', {'class': 'product__author-link'})
        author = author_elem.text.strip() if author_elem else ''

        price_elem = book.find('span', {'class': 'price-val'})
        price = price_elem.text.strip() if price_elem else ''

        if '-' in title:
            book_info = title.split('-', 1)
            worksheet.cell(row=i, column=1, value=book_info[0].strip())
            worksheet.cell(row=i, column=2, value=book_info[1].strip())
        else:
            worksheet.cell(row=i, column=1, value=title.strip())
            worksheet.cell(row=i, column=2, value=author.strip())

        worksheet.cell(row=i, column=3, value=price)

    for i in range(worksheet.max_row, 1, -1):
        if all([cell.value is None for cell in worksheet[i]]):
            worksheet.delete_rows(i, 1)

    workbook.save('books.xlsx')

